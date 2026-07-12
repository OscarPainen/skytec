"""Carga masiva de inventario desde .xlsx (plantilla, lectura, validación, import).

Toda la escritura en base pasa por `repo` (crear_producto / registrar_entrada /
set_disponible): aquí no hay SQL, solo Excel + validación. Los mensajes de error
están pensados para un usuario no técnico.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from modules.inventario import repo

# Encabezados exactos que el sistema espera (sin 'imagen': se agrega manual).
HEADERS = ["nombre", "descripcion", "categoria", "precio_venta", "costo",
           "stock_inicial", "disponible"]
OBLIGATORIAS = {"nombre", "precio_venta", "costo", "stock_inicial"}

_VERDADERO = {"si", "sí", "true", "verdadero", "1", "x"}
_FALSO = {"no", "false", "falso", "0"}

MSG_ILEGIBLE = (
    "No pudimos leer el archivo. Asegúrate de usar la plantilla descargable "
    "y de no cambiar los nombres de las columnas."
)
MSG_BLOQUEADO = (
    "El archivo está abierto en Excel. Ciérralo e inténtalo de nuevo."
)


class ExcelError(Exception):
    """Error a nivel de archivo, ya traducido a lenguaje humano."""


@dataclass
class FilaImport:
    numero_fila: int
    nombre: str
    descripcion: str
    categoria: str
    precio_venta: int
    costo: int
    stock_inicial: int
    disponible: bool
    error: str | None = None
    es_duplicado: bool = False

    @property
    def valida(self) -> bool:
        return self.error is None


# ── Generación de plantilla ─────────────────────────────────────────────────
def generar_plantilla(ruta: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    encabezado_fill = PatternFill("solid", fgColor="EFF4FE")
    encabezado_font = Font(bold=True, color="1E3A8A")
    for col, nombre in enumerate(HEADERS, start=1):
        celda = ws.cell(row=1, column=col, value=nombre)
        celda.font = encabezado_font
        celda.fill = encabezado_fill
        celda.alignment = Alignment(horizontal="center")

    ejemplos = [
        ["Cargador USB-C 20W", "Carga rápida original", "Accesorios", 12990, 6000, 15, "si"],
        ["Lámina templada iPhone 13", "Vidrio 9H", "Protección", 4990, 1500, 40, "si"],
    ]
    for fila in ejemplos:
        ws.append(fila)

    anchos = [26, 34, 16, 14, 10, 14, 12]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    inst = wb.create_sheet("Instrucciones")
    inst.column_dimensions["A"].width = 90
    notas = [
        "Instrucciones para la carga masiva de inventario — Skytec",
        "",
        "1. No cambies los nombres ni el orden de las columnas de la hoja 'Productos'.",
        "2. Columnas obligatorias: nombre, precio_venta, costo, stock_inicial.",
        "3. 'categoria' y 'descripcion' pueden quedar vacías.",
        "4. Números sin puntos de miles (ej: 12990, no 12.990). Usa punto para decimales.",
        "5. 'disponible': escribe 'si' o 'no' (por defecto 'si').",
        "6. La imagen del producto NO se carga aquí: se agrega después, manualmente,",
        "   desde el catálogo de cada producto.",
        "7. Borra las dos filas de ejemplo antes de cargar tus productos.",
    ]
    for i, texto in enumerate(notas, start=1):
        celda = inst.cell(row=i, column=1, value=texto)
        if i == 1:
            celda.font = Font(bold=True, size=13)

    wb.save(ruta)


# ── Lectura + validación ────────────────────────────────────────────────────
def _a_numero(valor) -> float:
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        raise ValueError("vacío")
    if isinstance(valor, str):
        valor = valor.strip().replace(",", ".")
    return float(valor)


def leer_archivo(ruta: str | Path) -> list[FilaImport]:
    """Lee y valida el .xlsx. Lanza ExcelError (humano) ante problemas de archivo."""
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    except PermissionError:
        raise ExcelError(MSG_BLOQUEADO)
    except (InvalidFileException, OSError, KeyError, Exception):
        raise ExcelError(MSG_ILEGIBLE)

    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        raise ExcelError("El archivo está vacío.")

    # Validar encabezados
    encabezados = [str(c).strip().lower() if c is not None else "" for c in filas[0]]
    encabezados = encabezados[: len(HEADERS)] if len(encabezados) >= len(HEADERS) else encabezados
    faltan = [h for h in HEADERS if h not in encabezados]
    if faltan:
        raise ExcelError(
            "El archivo no tiene el formato de la plantilla. "
            f"Faltan las columnas: {', '.join(faltan)}. "
            "Descarga la plantilla y no cambies los encabezados."
        )
    idx = {h: encabezados.index(h) for h in HEADERS}

    datos = filas[1:]
    if not any(any(c is not None and str(c).strip() for c in fila) for fila in datos):
        raise ExcelError("El archivo no tiene filas de datos para importar.")

    existentes = repo.mapa_nombres()
    vistos: set[str] = set()
    resultado: list[FilaImport] = []
    for i, fila in enumerate(datos, start=2):
        def val(col):
            j = idx[col]
            return fila[j] if j < len(fila) else None

        if not any(c is not None and str(c).strip() for c in fila):
            continue  # fila totalmente vacía: la ignoramos en silencio

        errores: list[str] = []
        nombre = str(val("nombre")).strip() if val("nombre") is not None else ""
        if not nombre:
            errores.append("falta el nombre")

        precio = costo = stock = 0
        try:
            precio = round(_a_numero(val("precio_venta")))
            if precio < 0:
                errores.append("precio_venta no puede ser negativo")
        except ValueError:
            errores.append("precio_venta debe ser un número")
        try:
            costo = round(_a_numero(val("costo")))
            if costo < 0:
                errores.append("costo no puede ser negativo")
        except ValueError:
            errores.append("costo debe ser un número")
        try:
            f_stock = _a_numero(val("stock_inicial"))
            if f_stock < 0 or f_stock != int(f_stock):
                errores.append("stock_inicial debe ser un entero ≥ 0")
            stock = int(f_stock)
        except ValueError:
            errores.append("stock_inicial debe ser un número entero")

        disp_raw = val("disponible")
        disp_txt = str(disp_raw).strip().lower() if disp_raw is not None else ""
        disponible = disp_txt not in _FALSO  # vacío o desconocido -> disponible

        clave = nombre.lower()
        es_dup = bool(nombre) and (clave in existentes or clave in vistos)
        if nombre:
            vistos.add(clave)

        resultado.append(FilaImport(
            numero_fila=i, nombre=nombre,
            descripcion=str(val("descripcion") or "").strip(),
            categoria=str(val("categoria") or "").strip(),
            precio_venta=precio, costo=costo, stock_inicial=stock,
            disponible=disponible,
            error="; ".join(errores) if errores else None,
            es_duplicado=es_dup,
        ))
    return resultado


# ── Importación (reutiliza repo, sin SQL propio) ────────────────────────────
def importar(filas: list[FilaImport], politica: str, usuario_id: int | None = None) -> dict:
    """politica: 'omitir' (default) o 'sumar' para duplicados. Devuelve reporte."""
    rep = {"importados": 0, "omitidos_error": 0, "omitidos_dup": 0, "sumados": 0}
    mapa = repo.mapa_nombres()
    for f in filas:
        if not f.valida:
            rep["omitidos_error"] += 1
            continue
        clave = f.nombre.lower()
        if clave in mapa:
            if politica == "sumar":
                if f.stock_inicial > 0:
                    repo.registrar_entrada(
                        mapa[clave], f.stock_inicial, f.costo, usuario_id,
                        motivo="Carga masiva (suma a existente)",
                    )
                rep["sumados"] += 1
            else:
                rep["omitidos_dup"] += 1
            continue
        pid = repo.crear_producto(
            nombre=f.nombre, precio_venta=f.precio_venta, costo=f.costo,
            stock_inicial=f.stock_inicial, categoria=f.categoria,
            descripcion=f.descripcion, imagen_path="", usuario_id=usuario_id,
        )
        if not f.disponible:
            repo.set_disponible(pid, False)
        mapa[clave] = pid
        rep["importados"] += 1
    return rep


if __name__ == "__main__":
    import os
    import tempfile
    from core import database

    tmp = Path(tempfile.mkdtemp())
    os.environ["SKYTEC_DB"] = str(tmp / "t.db")
    database.DB_PATH = tmp / "t.db"
    database.init_db()

    plantilla = tmp / "plantilla.xlsx"
    generar_plantilla(plantilla)
    filas = leer_archivo(plantilla)
    assert len(filas) == 2 and all(f.valida for f in filas), filas
    rep = importar(filas, "omitir")
    assert rep["importados"] == 2, rep
    # segunda pasada: ahora son duplicados
    filas2 = leer_archivo(plantilla)
    assert all(f.es_duplicado for f in filas2)
    assert importar(filas2, "omitir")["omitidos_dup"] == 2
    stock_antes = repo.listar_productos()[0].stock_actual
    importar(leer_archivo(plantilla), "sumar")
    assert repo.listar_productos(busqueda="Cargador")[0].stock_actual == stock_antes + 15

    # archivo con encabezados malos
    wb = openpyxl.Workbook(); wb.active.append(["foo", "bar"]); mal = tmp / "mal.xlsx"; wb.save(mal)
    try:
        leer_archivo(mal); raise AssertionError("debió fallar por encabezados")
    except ExcelError:
        pass

    # fila con errores de datos
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(HEADERS)
    ws.append(["", "d", "c", "abc", -5, 2.5, "si"])  # sin nombre, precio no num, costo neg, stock no entero
    bad = tmp / "bad.xlsx"; wb.save(bad)
    f = leer_archivo(bad)[0]
    assert not f.valida and "nombre" in f.error, f.error
    print("OK inventario/excel.py")
